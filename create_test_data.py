#!/usr/bin/env python3
"""
Test Excel file generator for Excel Translator App
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_test_excel():
    """Create a test Excel file with sample data."""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Japanese content sheet
    japanese_sheet = wb.create_sheet("Japanese Content")
    
    # Add headers with styling
    headers = ["項目", "説明", "詳細"]
    for col, header in enumerate(headers, 1):
        cell = japanese_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample Japanese content
    japanese_data = [
        ["製品名", "高品質な日本製品", "この製品は最新の技術を使用して製造されています"],
        ["価格", "¥50,000", "税込み価格です"],
        ["特徴", "軽量で持ち運びやすい", "重量はわずか500gです"],
        ["用途", "ビジネス用途に最適", "オフィスや会議室での使用に適しています"],
        ["保証", "2年間の品質保証", "製品に不具合がある場合は無料で交換いたします"],
        ["サポート", "24時間サポート", "お客様からのお問い合わせに24時間対応いたします"]
    ]
    
    for row, data in enumerate(japanese_data, 2):
        for col, value in enumerate(data, 1):
            cell = japanese_sheet.cell(row=row, column=col, value=value)
            if col == 1:  # First column - make it bold
                cell.font = Font(bold=True)
            if row % 2 == 0:  # Alternate row colors
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Auto-adjust column widths
    for column in japanese_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        japanese_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create Vietnamese content sheet
    vietnamese_sheet = wb.create_sheet("Vietnamese Content")
    
    # Add Vietnamese headers
    vn_headers = ["Mục", "Mô tả", "Chi tiết"]
    for col, header in enumerate(vn_headers, 1):
        cell = vietnamese_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="CC6600", end_color="CC6600", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample Vietnamese content
    vietnamese_data = [
        ["Tên sản phẩm", "Sản phẩm chất lượng cao", "Sản phẩm này được sản xuất bằng công nghệ tiên tiến"],
        ["Giá cả", "1.000.000 VNĐ", "Giá đã bao gồm thuế VAT"],
        ["Đặc điểm", "Nhẹ và dễ mang theo", "Trọng lượng chỉ 500g"],
        ["Công dụng", "Phù hợp cho công việc", "Thích hợp sử dụng trong văn phòng và phòng họp"],
        ["Bảo hành", "Bảo hành 2 năm", "Đổi mới miễn phí nếu có lỗi sản phẩm"],
        ["Hỗ trợ", "Hỗ trợ 24/7", "Đội ngũ hỗ trợ khách hàng 24 giờ"]
    ]
    
    for row, data in enumerate(vietnamese_data, 2):
        for col, value in enumerate(data, 1):
            cell = vietnamese_sheet.cell(row=row, column=col, value=value)
            if col == 1:  # First column - make it bold
                cell.font = Font(bold=True)
            if row % 2 == 0:  # Alternate row colors
                cell.fill = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in vietnamese_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        vietnamese_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create mixed content sheet
    mixed_sheet = wb.create_sheet("Mixed Content")
    
    # Add mixed content with formulas and numbers
    mixed_data = [
        ["Description", "Value", "Formula"],
        ["こんにちは世界", "=A2", "=LEN(A2)"],
        ["Price", 1000, "=B2*1.1"],
        ["Xin chào thế giới", "=A4", "=LEN(A4)"],
        ["Total", "=SUM(B2:B4)", "=AVERAGE(C2:C4)"],
        ["日本の文化は素晴らしい", 2023, "=YEAR(TODAY())"],
        ["Văn hóa Việt Nam rất đẹp", "=A7", "=LEN(A7)"]
    ]
    
    for row, data in enumerate(mixed_data, 1):
        for col, value in enumerate(data, 1):
            cell = mixed_sheet.cell(row=row, column=col, value=value)
            if row == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Auto-adjust column widths
    for column in mixed_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        mixed_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    output_path = Path("test_data.xlsx")
    wb.save(output_path)
    wb.close()
    
    print(f"Test Excel file created: {output_path}")
    print("The file contains:")
    print("- Japanese Content sheet with product information")
    print("- Vietnamese Content sheet with product information")
    print("- Mixed Content sheet with text, numbers, and formulas")
    print("\nYou can use this file to test the translator app!")

if __name__ == "__main__":
    try:
        create_test_excel()
    except ImportError:
        print("Error: openpyxl not installed. Run 'pip install openpyxl' first.")
    except Exception as e:
        print(f"Error creating test file: {e}")
