"""
Excel Handler Implementation

Infrastructure layer implementation for Excel file handling.
"""

import logging
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import asyncio
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import xlrd
import pandas as pd

from ...interfaces.repositories.file_repository_interface import FileRepositoryInterface
from ...domain.entities.excel_file import ExcelFile, SheetInfo, CellPosition
from ...domain.entities.formatting import (
    Formatting, CellFormat, FontFormat, BorderFormat, 
    AlignmentFormat, FillFormat, MergedCellInfo, RowFormat, ColumnFormat, SheetFormat
)


class ExcelHandler(FileRepositoryInterface):
    """Implementation of file repository for Excel files."""
    
    def __init__(self):
        """Initialize Excel handler."""
        self.logger = logging.getLogger(__name__)
    
    async def load_file(self, file_path: str) -> ExcelFile:
        """Load an Excel file and return file entity."""
        self.logger.info(f"Loading Excel file: {file_path}")
        
        try:
            excel_file = ExcelFile.create(file_path)
            
            if not excel_file.is_valid:
                return excel_file
            
            # Load workbook
            if excel_file.format_type in ['xlsx', 'xlsm']:
                workbook = openpyxl.load_workbook(file_path, data_only=False)
                await self._analyze_xlsx_file(workbook, excel_file)
            elif excel_file.format_type == 'xls':
                workbook = xlrd.open_workbook(file_path)
                await self._analyze_xls_file(workbook, excel_file)
            else:
                excel_file.is_valid = False
                excel_file.error_message = f"Unsupported format: {excel_file.format_type}"
            
            return excel_file
            
        except Exception as e:
            self.logger.error(f"Failed to load Excel file: {str(e)}", exc_info=True)
            excel_file = ExcelFile.create(file_path)
            excel_file.is_valid = False
            excel_file.error_message = str(e)
            return excel_file
    
    async def _analyze_xlsx_file(self, workbook: openpyxl.Workbook, excel_file: ExcelFile) -> None:
        """Analyze XLSX/XLSM file structure."""
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Count cells
            total_cells = 0
            text_cells = 0
            has_formulas = False
            has_merged_cells = len(worksheet.merged_cells.ranges) > 0
            
            # Get actual used range
            max_row = worksheet.max_row or 1
            max_col = worksheet.max_column or 1
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row, col)
                    total_cells += 1
                    
                    if cell.value is not None:
                        if isinstance(cell.value, str) and cell.value.strip():
                            text_cells += 1
                        elif str(cell.value).startswith('='):
                            has_formulas = True
            
            sheet_info = SheetInfo(
                name=sheet_name,
                text_cells=text_cells,
                total_cells=total_cells,
                has_formulas=has_formulas,
                has_merged_cells=has_merged_cells,
                row_count=max_row,
                column_count=max_col
            )
            
            excel_file.add_sheet_info(sheet_name, sheet_info)
    
    async def _analyze_xls_file(self, workbook: xlrd.Book, excel_file: ExcelFile) -> None:
        """Analyze XLS file structure."""
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            sheet_name = worksheet.name
            
            total_cells = worksheet.nrows * worksheet.ncols
            text_cells = 0
            has_formulas = False
            
            for row in range(worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell = worksheet.cell(row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        text_cells += 1
                    # XLS formula detection is limited
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        has_formulas = True
            
            sheet_info = SheetInfo(
                name=sheet_name,
                text_cells=text_cells,
                total_cells=total_cells,
                has_formulas=has_formulas,
                has_merged_cells=False,  # Merged cell detection for XLS is complex
                row_count=worksheet.nrows,
                column_count=worksheet.ncols
            )
            
            excel_file.add_sheet_info(sheet_name, sheet_info)
    
    async def save_file(self, excel_file: ExcelFile, output_path: str) -> bool:
        """Save Excel file to specified path."""
        try:
            # This would involve creating a new workbook with the file data
            # For now, return True as placeholder
            self.logger.info(f"Saving Excel file to: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to save Excel file: {str(e)}")
            return False
    
    async def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names from Excel file."""
        try:
            if file_path.endswith(('.xlsx', '.xlsm')):
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                return workbook.sheetnames
            elif file_path.endswith('.xls'):
                workbook = xlrd.open_workbook(file_path)
                return workbook.sheet_names()
            else:
                return []
                
        except Exception as e:
            self.logger.error(f"Failed to get sheet names: {str(e)}")
            return []
    
    async def get_sheet_info(self, file_path: str, sheet_name: str) -> SheetInfo:
        """Get detailed information about a specific sheet."""
        try:
            if file_path.endswith(('.xlsx', '.xlsm')):
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                worksheet = workbook[sheet_name]
                
                text_cells = 0
                total_cells = 0
                has_formulas = False
                has_merged_cells = len(worksheet.merged_cells.ranges) > 0
                
                max_row = worksheet.max_row or 1
                max_col = worksheet.max_column or 1
                
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row, col)
                        total_cells += 1
                        
                        if cell.value is not None:
                            if isinstance(cell.value, str) and cell.value.strip():
                                text_cells += 1
                            elif str(cell.value).startswith('='):
                                has_formulas = True
                
                return SheetInfo(
                    name=sheet_name,
                    text_cells=text_cells,
                    total_cells=total_cells,
                    has_formulas=has_formulas,
                    has_merged_cells=has_merged_cells,
                    row_count=max_row,
                    column_count=max_col
                )
            
            # Fallback for unsupported formats
            return SheetInfo(
                name=sheet_name,
                text_cells=0,
                total_cells=0,
                has_formulas=False,
                has_merged_cells=False,
                row_count=0,
                column_count=0
            )
            
        except Exception as e:
            self.logger.error(f"Failed to get sheet info: {str(e)}")
            return SheetInfo(
                name=sheet_name,
                text_cells=0,
                total_cells=0,
                has_formulas=False,
                has_merged_cells=False,
                row_count=0,
                column_count=0
            )
    
    async def extract_text_content(self, file_path: str, sheet_name: str) -> List[tuple]:
        """Extract text content from sheet as list of (text, row, column) tuples."""
        try:
            content = []
            
            if file_path.endswith(('.xlsx', '.xlsm')):
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                worksheet = workbook[sheet_name]
                
                max_row = worksheet.max_row or 1
                max_col = worksheet.max_column or 1
                
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row, col)
                        
                        if (cell.value is not None and 
                            isinstance(cell.value, str) and 
                            cell.value.strip()):
                            content.append((cell.value.strip(), row, col))
            
            elif file_path.endswith('.xls'):
                workbook = xlrd.open_workbook(file_path)
                worksheet = workbook.sheet_by_name(sheet_name)
                
                for row in range(worksheet.nrows):
                    for col in range(worksheet.ncols):
                        cell_value = worksheet.cell(row, col).value
                        
                        if (cell_value and 
                            isinstance(cell_value, str) and 
                            cell_value.strip()):
                            content.append((cell_value.strip(), row + 1, col + 1))
            
            self.logger.info(f"Extracted {len(content)} text cells from {sheet_name}")
            return content
            
        except Exception as e:
            self.logger.error(f"Failed to extract text content: {str(e)}")
            return []
    
    async def extract_formatting(self, file_path: str, sheet_name: str) -> Formatting:
        """Extract formatting information from sheet."""
        try:
            formatting = Formatting.create_empty(sheet_name)
            
            if file_path.endswith(('.xlsx', '.xlsm')):
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook[sheet_name]
                
                # Extract merged cells
                for merged_range in worksheet.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = merged_range.bounds
                    merged_info = MergedCellInfo(min_row, min_col, max_row, max_col)
                    formatting.add_merged_cell(merged_info)
                
                # Extract cell formatting (sample implementation)
                max_row = min(worksheet.max_row or 1, 100)  # Limit for performance
                max_col = min(worksheet.max_column or 1, 50)
                
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row, col)
                        
                        if cell.value is not None:
                            # Extract basic formatting
                            font_format = FontFormat(
                                name=cell.font.name,
                                size=cell.font.size,
                                color=cell.font.color.rgb if cell.font.color else None
                            )
                            
                            cell_format = CellFormat(
                                font=font_format,
                                border=BorderFormat(),
                                alignment=AlignmentFormat(),
                                fill=FillFormat()
                            )
                            
                            formatting.set_cell_format(row, col, cell_format)
            
            return formatting
            
        except Exception as e:
            self.logger.error(f"Failed to extract formatting: {str(e)}")
            return Formatting.create_empty(sheet_name)
    
    async def apply_translations(self, file_path: str, translations: Dict[str, List[tuple]], 
                                output_path: str) -> bool:
        """Apply translations to file and save as new file."""
        try:
            self.logger.info(f"Applying translations from {file_path} to {output_path}")
            
            if file_path.endswith(('.xlsx', '.xlsm')):
                # Load original workbook
                workbook = openpyxl.load_workbook(file_path)
                
                # Apply translations to each sheet
                for sheet_name, sheet_translations in translations.items():
                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        for translated_text, row, col in sheet_translations:
                            cell = worksheet.cell(row, col)
                            cell.value = translated_text
                
                # Save as new file
                workbook.save(output_path)
                return True
            
            elif file_path.endswith('.xls'):
                # XLS files need to be converted to XLSX for writing
                # Load with pandas and save as XLSX
                with pd.ExcelFile(file_path) as xls:
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        for sheet_name in xls.sheet_names:
                            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                            
                            # Apply translations if available
                            if sheet_name in translations:
                                for translated_text, row, col in translations[sheet_name]:
                                    if row-1 < len(df) and col-1 < len(df.columns):
                                        df.iloc[row-1, col-1] = translated_text
                            
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                return True
            
            return False
            
        except Exception as e:
            self.logger.error(f"Failed to apply translations: {str(e)}", exc_info=True)
            return False
    
    async def validate_file(self, file_path: str) -> Dict[str, Any]:
        """Validate if file can be processed."""
        try:
            path = Path(file_path)
            
            validation = {
                'is_valid': True,
                'is_supported_format': True,
                'is_readable': True,
                'is_corrupted': False,
                'has_password_protection': False,
                'file_size_mb': 0.0,
                'warnings': [],
                'errors': []
            }
            
            # Check file existence
            if not path.exists():
                validation['is_valid'] = False
                validation['errors'].append("File does not exist")
                return validation
            
            # Check file format
            if path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
                validation['is_supported_format'] = False
                validation['errors'].append(f"Unsupported format: {path.suffix}")
            
            # Check file size
            file_size = path.stat().st_size
            validation['file_size_mb'] = file_size / (1024 * 1024)
            
            if validation['file_size_mb'] > 100:
                validation['warnings'].append(f"Large file ({validation['file_size_mb']:.1f} MB)")
            
            # Try to open file to check for corruption/password protection
            try:
                if path.suffix.lower() in ['.xlsx', '.xlsm']:
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    workbook.close()
                elif path.suffix.lower() == '.xls':
                    workbook = xlrd.open_workbook(file_path)
            except Exception as e:
                error_msg = str(e).lower()
                if 'password' in error_msg or 'encrypted' in error_msg:
                    validation['has_password_protection'] = True
                    validation['errors'].append("File is password protected")
                else:
                    validation['is_corrupted'] = True
                    validation['errors'].append("File appears to be corrupted")
                validation['is_valid'] = False
            
            return validation
            
        except Exception as e:
            return {
                'is_valid': False,
                'is_supported_format': False,
                'is_readable': False,
                'is_corrupted': True,
                'has_password_protection': False,
                'file_size_mb': 0.0,
                'warnings': [],
                'errors': [f"Validation failed: {str(e)}"]
            }
    
    async def backup_file(self, file_path: str, backup_path: str) -> bool:
        """Create backup of original file."""
        try:
            import shutil
            shutil.copy2(file_path, backup_path)
            self.logger.info(f"Backup created: {backup_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to create backup: {str(e)}")
            return False
    
    async def get_file_metadata(self, file_path: str) -> Dict[str, Any]:
        """Get file metadata including size, dates, etc."""
        try:
            path = Path(file_path)
            
            if not path.exists():
                return {}
            
            stat = path.stat()
            
            return {
                'file_name': path.name,
                'file_size': stat.st_size,
                'file_size_mb': stat.st_size / (1024 * 1024),
                'created_date': datetime.fromtimestamp(stat.st_ctime),
                'modified_date': datetime.fromtimestamp(stat.st_mtime),
                'accessed_date': datetime.fromtimestamp(stat.st_atime),
                'format': path.suffix.lower().replace('.', ''),
                'is_file': path.is_file(),
                'is_readable': path.stat().st_mode & 0o444,
                'absolute_path': str(path.absolute())
            }
            
        except Exception as e:
            self.logger.error(f"Failed to get file metadata: {str(e)}")
            return {}
