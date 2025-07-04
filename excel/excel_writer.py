import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional
import logging
import os
from copy import copy

class ExcelWriter:
    """Handles writing translated content to Excel files while preserving formatting."""
    
    def __init__(self, source_file_path: str, output_file_path: str):
        """Initialize Excel writer.
        
        Args:
            source_file_path: Path to the source Excel file
            output_file_path: Path for the output Excel file
        """
        self.source_file_path = source_file_path
        self.output_file_path = output_file_path
        self.source_workbook = None
        self.output_workbook = None
        self.logger = logging.getLogger(__name__)
    
    def create_translated_workbook(self, translation_data: Dict[str, List[Tuple[str, int, int]]]):
        """Create a new workbook with translated content.
        
        Args:
            translation_data: Dictionary mapping sheet names to lists of (translated_text, row, col)
        """
        try:
            # Load the source workbook
            self.source_workbook = openpyxl.load_workbook(self.source_file_path)
            
            # Create output workbook by copying the source
            self.output_workbook = Workbook()
            
            # Remove the default sheet
            if 'Sheet' in self.output_workbook.sheetnames:
                self.output_workbook.remove(self.output_workbook['Sheet'])
            
            # Process each sheet
            for sheet_name in self.source_workbook.sheetnames:
                self._copy_sheet_with_translations(sheet_name, translation_data.get(sheet_name, []))
            
            # Save the output workbook
            self.output_workbook.save(self.output_file_path)
            self.logger.info(f"Translated workbook saved to: {self.output_file_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to create translated workbook: {str(e)}")
            raise
        finally:
            self._cleanup()
    
    def _copy_sheet_with_translations(self, sheet_name: str, translations: List[Tuple[str, int, int]]):
        """Copy a sheet with translated content.
        
        Args:
            sheet_name: Name of the sheet to copy
            translations: List of (translated_text, row, col) tuples
        """
        source_sheet = self.source_workbook[sheet_name]
        target_sheet = self.output_workbook.create_sheet(title=sheet_name)
        
        # Create a mapping of (row, col) to translated text
        translation_map = {(row, col): text for text, row, col in translations}
        
        # Copy all cells with their formatting
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # Use translated text if available, otherwise use original value
                cell_key = (cell.row, cell.column)
                if cell_key in translation_map:
                    target_cell.value = translation_map[cell_key]
                else:
                    target_cell.value = cell.value
                
                # Copy cell formatting
                self._copy_cell_style(cell, target_cell)
        
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        
        # Copy column dimensions
        for col_letter, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = dimension.width
            target_sheet.column_dimensions[col_letter].hidden = dimension.hidden
        
        # Copy row dimensions
        for row_num, dimension in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_num].height = dimension.height
            target_sheet.row_dimensions[row_num].hidden = dimension.hidden
        
        # Copy sheet properties
        target_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor
        target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
        target_sheet.sheet_view.showRowColHeaders = source_sheet.sheet_view.showRowColHeaders
        
        # Copy print settings
        target_sheet.page_setup = copy(source_sheet.page_setup)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        
        self.logger.info(f"Copied sheet '{sheet_name}' with {len(translations)} translations")
    
    def _copy_cell_style(self, source_cell, target_cell):
        """Copy formatting from source cell to target cell.
        
        Args:
            source_cell: Source cell with formatting
            target_cell: Target cell to apply formatting to
        """
        try:
            # Copy font
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            
            # Copy fill
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            
            # Copy border
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            
            # Copy number format
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            
            # Copy protection
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            
        except Exception as e:
            self.logger.warning(f"Failed to copy cell style: {str(e)}")
    
    def _cleanup(self):
        """Clean up resources."""
        if self.source_workbook:
            self.source_workbook.close()
            self.source_workbook = None
        
        if self.output_workbook:
            self.output_workbook.close()
            self.output_workbook = None
    
    @staticmethod
    def generate_output_filename(source_file_path: str, suffix: str = "_translated") -> str:
        """Generate output filename based on source file.
        
        Args:
            source_file_path: Path to the source file
            suffix: Suffix to add to the filename
            
        Returns:
            Generated output file path
        """
        base_name = os.path.splitext(source_file_path)[0]
        extension = os.path.splitext(source_file_path)[1]
        return f"{base_name}{suffix}{extension}"
    
    def validate_output_file(self) -> bool:
        """Validate that the output file was created successfully.
        
        Returns:
            True if the output file exists and is valid
        """
        if not os.path.exists(self.output_file_path):
            return False
        
        try:
            # Try to open the file to verify it's valid
            test_workbook = openpyxl.load_workbook(self.output_file_path)
            test_workbook.close()
            return True
        except Exception as e:
            self.logger.error(f"Output file validation failed: {str(e)}")
            return False
