import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional
import logging
import re

class ExcelReader:
    """Handles reading Excel files and extracting translatable content."""
    
    def __init__(self, file_path: str):
        """Initialize Excel reader.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.logger = logging.getLogger(__name__)
        self.load_workbook()
    
    def load_workbook(self):
        """Load the Excel workbook."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            self.logger.info(f"Loaded workbook: {self.file_path}")
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {str(e)}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook."""
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def get_sheet_info(self) -> Dict[str, Dict[str, Any]]:
        """Get information about all sheets."""
        if not self.workbook:
            return {}
        
        sheet_info = {}
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Count cells with content
            cell_count = 0
            text_cell_count = 0
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_count += 1
                        if self._is_translatable_text(cell.value):
                            text_cell_count += 1
            
            sheet_info[sheet_name] = {
                "total_cells": cell_count,
                "text_cells": text_cell_count,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            }
        
        return sheet_info
    
    def extract_translatable_content(self, sheet_name: str) -> Tuple[List[str], List[Tuple[int, int]]]:
        """Extract translatable content from a sheet.
        
        Args:
            sheet_name: Name of the sheet to extract content from
            
        Returns:
            Tuple of (texts, cell_positions)
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return [], []
        
        sheet = self.workbook[sheet_name]
        texts = []
        positions = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and self._is_translatable_text(cell.value):
                    texts.append(str(cell.value))
                    positions.append((cell.row, cell.column))
        
        self.logger.info(f"Extracted {len(texts)} translatable texts from sheet '{sheet_name}'")
        return texts, positions
    
    def _is_translatable_text(self, value: Any) -> bool:
        """Check if a cell value is translatable text.
        
        Args:
            value: Cell value to check
            
        Returns:
            True if the value is translatable text
        """
        if not isinstance(value, str):
            return False
        
        # Skip empty or whitespace-only strings
        if not value.strip():
            return False
        
        # Skip if it's purely numeric
        if value.strip().replace('.', '').replace(',', '').replace('-', '').isdigit():
            return False
        
        # Skip if it's a formula (starts with =)
        if value.strip().startswith('='):
            return False
        
        # Skip if it's a date pattern
        date_patterns = [
            r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$',
            r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',
            r'^\d{1,2}:\d{2}(:\d{2})?$'
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return False
        
        # Skip if it's mostly numbers with some formatting
        if re.match(r'^[\d\s.,%-]+$', value.strip()):
            return False
        
        # Skip common non-translatable patterns
        non_translatable_patterns = [
            r'^[A-Z]{1,10}\d+$',  # Cell references like A1, B2C3
            r'^#[A-Z]+!?$',       # Error values like #REF!, #NAME?
            r'^[A-Z_]+$',         # All caps constants
            r'^\$[A-Z]+\$\d+$',   # Absolute references
        ]
        
        for pattern in non_translatable_patterns:
            if re.match(pattern, value.strip()):
                return False
        
        return True
    
    def get_cell_style_info(self, sheet_name: str, row: int, col: int) -> Dict[str, Any]:
        """Get style information for a specific cell.
        
        Args:
            sheet_name: Name of the sheet
            row: Row number (1-based)
            col: Column number (1-based)
            
        Returns:
            Dictionary containing style information
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {}
        
        sheet = self.workbook[sheet_name]
        cell = sheet.cell(row=row, column=col)
        
        style_info = {
            'font': {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'color': cell.font.color.rgb if cell.font.color else None
            },
            'fill': {
                'fill_type': cell.fill.fill_type,
                'start_color': cell.fill.start_color.rgb if cell.fill.start_color else None,
                'end_color': cell.fill.end_color.rgb if cell.fill.end_color else None
            },
            'border': {
                'left': str(cell.border.left.style) if cell.border.left else None,
                'right': str(cell.border.right.style) if cell.border.right else None,
                'top': str(cell.border.top.style) if cell.border.top else None,
                'bottom': str(cell.border.bottom.style) if cell.border.bottom else None
            },
            'alignment': {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'text_rotation': cell.alignment.text_rotation,
                'wrap_text': cell.alignment.wrap_text
            },
            'number_format': cell.number_format
        }
        
        return style_info
    
    def get_merged_cells(self, sheet_name: str) -> List[str]:
        """Get merged cell ranges for a sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            List of merged cell range strings
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        
        sheet = self.workbook[sheet_name]
        return [str(merged_range) for merged_range in sheet.merged_cells.ranges]
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
